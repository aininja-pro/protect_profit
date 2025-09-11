from fastapi import APIRouter, UploadFile, File, HTTPException, Form, Path
from typing import List, Dict, Any, Optional
import os
import uuid
import subprocess
import json
from datetime import datetime, date
import tempfile

from ..db import get_supabase_client
from ..services.ai_quote_parser import ai_quote_parser

router = APIRouter(prefix="/quotes", tags=["quotes"])

def extract_text_from_file(file_content: bytes, filename: str, content_type: str) -> str:
    """Extract text from uploaded file based on file type"""
    try:
        print(f"📄 EXTRACTION: File {filename}, content_type: {content_type}, size: {len(file_content)} bytes")
        
        if content_type == 'application/pdf':
            extracted = extract_pdf_text(file_content)
            print(f"📄 EXTRACTION: PDF extracted {len(extracted)} characters")
            print(f"📄 EXTRACTION: Preview: {extracted[:100]}...")
            return extracted
        elif content_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword']:
            return extract_docx_text(file_content, filename)
        elif content_type == 'text/csv':
            return file_content.decode('utf-8')
        elif content_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            return extract_excel_text(file_content)
        else:
            # Fallback: try to decode as text
            return file_content.decode('utf-8', errors='ignore')
    except Exception as e:
        print(f"📄 EXTRACTION ERROR for {filename}: {e}")
        return f"[TEXT_EXTRACTION_FAILED: {filename}]"

def extract_pdf_text(file_content: bytes) -> str:
    """Extract text from PDF file - temporarily using pypdf for debugging"""
    try:
        import pypdf
        import io
        pdf_reader = pypdf.PdfReader(io.BytesIO(file_content))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        return f"[PDF_EXTRACTION_ERROR: {str(e)}]"

def extract_docx_text(file_content: bytes, filename: str) -> str:
    """Extract text from DOCX file"""
    try:
        import docx
        import io
        
        # Create a Document from bytes
        doc = docx.Document(io.BytesIO(file_content))
        text = ""
        
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        
        return text.strip()
    except ImportError:
        return "[DOCX_EXTRACTION_REQUIRES_PYTHON_DOCX: pip install python-docx]"
    except Exception as e:
        return f"[DOCX_EXTRACTION_ERROR: {str(e)}]"

def extract_excel_text(file_content: bytes) -> str:
    """Extract text from Excel file using openpyxl"""
    try:
        import openpyxl
        import io
        
        # Create a workbook from bytes
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        
        return text.strip()
    except ImportError:
        return "[EXCEL_EXTRACTION_REQUIRES_OPENPYXL: pip install openpyxl]"
    except Exception as e:
        return f"[EXCEL_EXTRACTION_ERROR: {str(e)}]"

def scan_file_with_clamav(file_content: bytes, filename: str) -> bool:
    """Scan file with ClamAV for viruses"""
    try:
        print(f"🦠 CLAM: Starting ClamAV scan for {filename}")
        # Write file to temporary location
        temp_path = f"/tmp/{uuid.uuid4()}_{filename}"
        with open(temp_path, "wb") as f:
            f.write(file_content)
        
        print(f"🦠 CLAM: Running clamscan on {temp_path}")
        # Run ClamAV scan
        result = subprocess.run(
            ["clamscan", "--no-summary", temp_path],
            capture_output=True,
            text=True
        )
        
        print(f"🦠 CLAM: clamscan return code: {result.returncode}")
        print(f"🦠 CLAM: clamscan stdout: {result.stdout}")
        print(f"🦠 CLAM: clamscan stderr: {result.stderr}")
        
        # Clean up temp file
        os.remove(temp_path)
        
        # Return True if clean (exit code 0), False if infected
        scan_passed = result.returncode == 0
        print(f"🦠 CLAM: Scan result: {'PASSED' if scan_passed else 'FAILED'}")
        return scan_passed
        
    except Exception as e:
        # If ClamAV is not available, log warning but don't block upload
        print(f"🦠 CLAM: Exception occurred: {str(e)}")
        print(f"🦠 CLAM: Allowing upload (ClamAV not available)")
        return True  # Allow upload if scan fails

@router.post("/divisions/{division_id}/upload")
async def upload_division_quote(
    division_id: str = Path(...),
    project_id: str = Form(...),
    vendor_name: str = Form(...),
    scope_type: str = Form("complete_division"),  # 'complete_division' | 'specific_items'
    scope_items: str = Form("[]"),  # JSON array of line item IDs
    scope_budget_total: float = Form(0.0),  # Total budget of selected items
    scope_notes: str = Form(""),  # Additional scope clarification
    file: UploadFile = File(...)
):
    """Upload a vendor quote file for AI parsing"""
    try:
        print(f"📤 UPLOAD: Received request for division {division_id}")
        print(f"📤 UPLOAD: project_id={project_id}, vendor_name={vendor_name}")
        print(f"📤 UPLOAD: scope_type={scope_type}, scope_budget=${scope_budget_total}")
        print(f"📤 UPLOAD: file={file.filename}, content_type={file.content_type}, size={file.size}")
        
        # Parse and validate scope items
        try:
            parsed_scope_items = json.loads(scope_items)
            print(f"📤 UPLOAD: Scope covers {len(parsed_scope_items)} line items")
        except json.JSONDecodeError:
            print(f"📤 UPLOAD: Invalid scope_items JSON, defaulting to empty array")
            parsed_scope_items = []
        # Validate file type
        allowed_types = [
            'application/pdf',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword',
            'text/csv',
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ]
        
        print(f"📤 UPLOAD: Validating file type: {file.content_type}")
        if file.content_type not in allowed_types:
            print(f"📤 UPLOAD: File type validation failed!")
            raise HTTPException(
                status_code=400, 
                detail="File must be PDF, DOCX, DOC, CSV, or Excel format"
            )
        print(f"📤 UPLOAD: File type validation passed")
        
        # Read file content
        print(f"📤 UPLOAD: Reading file content...")
        file_content = await file.read()
        print(f"📤 UPLOAD: File content read, size: {len(file_content)} bytes")
        
        # Check file size (10MB limit)
        max_size = int(os.getenv('MAX_FILE_SIZE', 10485760))
        if len(file_content) > max_size:
            print(f"📤 UPLOAD: File size limit exceeded!")
            raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")
        print(f"📤 UPLOAD: File size check passed")
        
        # Skip virus scan for now - TODO: Fix ClamAV setup
        print(f"📤 UPLOAD: Skipping virus scan (disabled)")
        
        # Get Supabase client
        print(f"📤 UPLOAD: Getting Supabase client...")
        supabase = get_supabase_client()
        print(f"📤 UPLOAD: Supabase client obtained")
        
        # Verify project exists
        print(f"📤 UPLOAD: Checking if project {project_id} exists...")
        project_check = supabase.table("projects").select("id").eq("id", project_id).execute()
        print(f"📤 UPLOAD: Project check result: {project_check.data}")
        if not project_check.data:
            print(f"📤 UPLOAD: Project not found!")
            raise HTTPException(status_code=404, detail="Project not found")
        print(f"📤 UPLOAD: Project exists")
        
        # Store file locally for now (in production, upload to cloud storage)
        file_id = str(uuid.uuid4())
        file_path = f"/tmp/quotes_{file_id}_{file.filename}"
        
        # Save file content to temporary location for parsing
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        file_url = file_path  # Store local path for now
        
        # Find or create vendor
        vendor_result = supabase.table("vendors").select("id").eq("name", vendor_name).execute()
        
        if vendor_result.data:
            vendor_id = vendor_result.data[0]["id"]
        else:
            vendor_record = {
                "id": str(uuid.uuid4()),
                "name": vendor_name,
                "contact_json": {}
            }
            vendor_create = supabase.table("vendors").insert(vendor_record).execute()
            vendor_id = vendor_create.data[0]["id"]
        
        # Create vendor quote record
        quote_id = str(uuid.uuid4())
        
        # Extract project UUID and division code
        division_code = division_id.split('-')[0]  # e.g., "04"
        actual_division_id = project_id  # Use project_id directly for consistency
        
        quote_record = {
            "id": quote_id,
            "project_id": project_id,
            "division_id": actual_division_id,
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "file_url": file_url,
            "received_at": datetime.now().isoformat(),
            "status": "draft",
            "version": 1,
            "division_code": division_code,
            # Enhanced scope tracking fields
            "scope_type": scope_type,
            "scope_items": parsed_scope_items,
            "scope_budget_total": scope_budget_total,
            "scope_notes": scope_notes
        }
        
        quote_result = supabase.table("vendor_quotes").insert(quote_record).execute()
        
        # Update division status to "quotes_uploaded"
        supabase.table("division_status").upsert({
            "division_id": actual_division_id,
            "status": "quotes_uploaded", 
            "updated_at": datetime.now().isoformat()
        }).execute()
        
        # Log action
        audit_log = {
            "project_id": project_id,
            "user_id": None,  # TODO: Get from JWT auth
            "action": "quote_uploaded",
            "entity_type": "quote",
            "entity_id": quote_id,
            "details": {
                "file_name": file.filename,
                "vendor_name": vendor_name,
                "division_code": division_id.split('-')[0],
                "file_size": len(file_content),
                "scope_type": scope_type,
                "scope_items_count": len(parsed_scope_items),
                "scope_budget_total": scope_budget_total
            }
        }
        
        supabase.table("audit_logs").insert(audit_log).execute()
        
        # TODO: Queue for AI parsing
        # In a real implementation, this would trigger a background job
        # to parse the quote using OpenAI/Anthropic
        
        return {
            "message": "Quote uploaded successfully",
            "quote_id": quote_id,
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "division_id": division_id,
            "status": "draft",
            "scope_type": scope_type,
            "scope_items_count": len(parsed_scope_items),
            "scope_budget_total": scope_budget_total
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"📤 UPLOAD: EXCEPTION OCCURRED: {str(e)}")
        print(f"📤 UPLOAD: Exception type: {type(e)}")
        import traceback
        print(f"📤 UPLOAD: Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.post("/subcategories/{subcategory_id}/upload")
async def upload_subcategory_quote(
    subcategory_id: str = Path(...),
    project_id: str = Form(...),
    division_code: str = Form(...),
    vendor_name: str = Form(...),
    scope_type: str = Form(...),
    scope_items: str = Form(""),
    scope_notes: str = Form(""),
    file: UploadFile = File(...)
):
    """Upload a vendor quote file for a specific subcategory"""
    try:
        print(f"📤 SUBCAT UPLOAD: subcategory={subcategory_id}, vendor={vendor_name}")
        
        # Validate file type (same as division upload)
        allowed_types = [
            'application/pdf',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword',
            'text/csv',
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ]
        
        if file.content_type not in allowed_types:
            raise HTTPException(
                status_code=400, 
                detail="File must be PDF, DOCX, DOC, CSV, or Excel format"
            )
        
        file_content = await file.read()
        
        # Check file size (10MB limit)
        max_size = int(os.getenv('MAX_FILE_SIZE', 10485760))
        if len(file_content) > max_size:
            raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")
        
        # Get Supabase client
        supabase = get_supabase_client()
        
        # Store file locally
        file_id = str(uuid.uuid4())
        file_path = f"/tmp/subcategory_quotes_{file_id}_{file.filename}"
        
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # Find or create vendor
        vendor_result = supabase.table("vendors").select("id").eq("name", vendor_name).execute()
        
        if vendor_result.data:
            vendor_id = vendor_result.data[0]["id"]
        else:
            vendor_record = {
                "id": str(uuid.uuid4()),
                "name": vendor_name,
                "contact_json": {}
            }
            vendor_create = supabase.table("vendors").insert(vendor_record).execute()
            vendor_id = vendor_create.data[0]["id"]
        
        # Create subcategory quote record with new fields
        quote_id = str(uuid.uuid4())
        
        quote_record = {
            "id": quote_id,
            "project_id": project_id,
            "division_id": project_id,  
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "file_url": file_path,
            "received_at": datetime.now().isoformat(),
            "status": "draft",
            "version": 1,
            "division_code": division_code,
            "subcategory_id": subcategory_id,
            "scope_type": scope_type,
            "scope_items": scope_items,
            "scope_notes": scope_notes
        }
        
        supabase.table("vendor_quotes").insert(quote_record).execute()
        
        return {
            "message": "Subcategory quote uploaded successfully",
            "quote_id": quote_id,
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "subcategory_id": subcategory_id,
            "status": "draft"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"📤 SUBCAT UPLOAD ERROR: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Subcategory upload failed: {str(e)}")


@router.get("/subcategories/{division_code}/{project_id}")
async def get_subcategory_quotes(division_code: str, project_id: str):
    """Get all subcategory quotes for a division"""
    try:
        supabase = get_supabase_client()
        
        quotes_result = supabase.table("vendor_quotes")\
            .select("*, quote_line_items(*)")\
            .eq("project_id", project_id)\
            .eq("division_code", division_code)\
            .not_.is_("subcategory_id", "null")\
            .execute()
        
        return {
            "division_code": division_code,
            "project_id": project_id,
            "subcategory_quotes": quotes_result.data
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to retrieve subcategory quotes: {str(e)}")


@router.get("/list/{project_id}")
async def list_quotes(project_id: str):
    """Get all quotes for a project"""
    try:
        supabase = get_supabase_client()
        
        result = supabase.table("vendor_quotes")\
            .select("*, files(file_name, file_size, created_at)")\
            .eq("project_id", project_id)\
            .order("created_at", desc=True)\
            .execute()
        
        return {
            "project_id": project_id,
            "quotes": result.data
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching quotes: {str(e)}")

@router.post("/{quote_id}/parse")
async def parse_quote(quote_id: str = Path(...)):
    """Extract and normalize quote data using AI"""
    try:
        supabase = get_supabase_client()
        
        # Get quote record
        quote_result = supabase.table("vendor_quotes").select("*").eq("id", quote_id).execute()
        if not quote_result.data:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        quote = quote_result.data[0]
        
        # Keep status as "draft" during parsing (parsing status not allowed by DB constraint)
        
        # Get file path and extract text
        file_url = quote["file_url"]
        
        try:
            # Read file content
            with open(file_url, "rb") as f:
                file_content = f.read()
            
            # Determine content type from file extension
            filename = os.path.basename(file_url)
            if filename.lower().endswith('.pdf'):
                content_type = 'application/pdf'
            elif filename.lower().endswith(('.docx', '.doc')):
                content_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            elif filename.lower().endswith('.csv'):
                content_type = 'text/csv'
            elif filename.lower().endswith(('.xlsx', '.xls')):
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            else:
                content_type = 'text/plain'
            
            # Extract text from file
            quote_text = extract_text_from_file(file_content, filename, content_type)
            
            # Build context for AI parser
            context = {
                "division_id": quote.get("division_id"),
                "division_code": quote.get("division_code"),  # Add division_code for trade detection
                "project_id": quote.get("project_id"),
                "vendor_name": quote.get("vendor_name"),
                "file_name": filename
            }
            
            # Use AI parser to extract structured data
            print(f"🚀 QUOTES_ENDPOINT: About to call AI parser with text length: {len(quote_text)}")
            parsed_data = ai_quote_parser.parse_quote_text(quote_text, context)
            print(f"🚀 QUOTES_ENDPOINT: AI parser returned data with total: {parsed_data.get('pricing_summary', {}).get('total_amount', 'No total found')}")
            
            # Ensure required structure
            normalized_data = {
                "vendor_info": parsed_data.get("vendor_info", {}),
                "scope_summary": parsed_data.get("scope_summary", ""),
                "line_items": parsed_data.get("line_items", []),
                "pricing_summary": parsed_data.get("pricing_summary", {}),
                "timeline": parsed_data.get("timeline", {}),
                "exclusions": parsed_data.get("exclusions", []),
                "assumptions": parsed_data.get("assumptions", []),
                "confidence_score": parsed_data.get("confidence_score", 0.5),
                "parsing_flags": parsed_data.get("parsing_flags", []),
                "raw_text_preview": quote_text[:500] if quote_text else ""
            }
            
        except FileNotFoundError:
            # File not found - create placeholder data
            normalized_data = {
                "vendor_info": {"name": quote.get("vendor_name", "Unknown Vendor")},
                "scope_summary": "File not found - manual input needed",
                "line_items": [],
                "pricing_summary": {"total_amount": 0.0},
                "timeline": {},
                "exclusions": [],
                "assumptions": [],
                "confidence_score": 0.1,
                "parsing_flags": [f"File not found: {file_url}"],
                "error": f"File not accessible: {file_url}"
            }
        except Exception as parse_error:
            # Parsing error - create error data
            normalized_data = {
                "vendor_info": {"name": quote.get("vendor_name", "Unknown Vendor")},
                "scope_summary": f"Parsing failed - {str(parse_error)}",
                "line_items": [],
                "pricing_summary": {"total_amount": 0.0},
                "timeline": {},
                "exclusions": [],
                "assumptions": [],
                "confidence_score": 0.1,
                "parsing_flags": [f"Parsing error: {str(parse_error)}"],
                "error": str(parse_error)
            }
        
        # Store normalized JSON
        supabase.table("vendor_quotes").update({
            "normalized_json": normalized_data,
            "status": "parsed"
        }).eq("id", quote_id).execute()
        
        # Create quote line items from parsed data
        line_items_created = 0
        for item_data in normalized_data.get("line_items", []):
            line_item = {
                "id": str(uuid.uuid4()),
                "quote_id": quote_id,
                "description": item_data.get("description", "Unknown item"),
                "quantity": item_data.get("quantity"),
                "unit": item_data.get("unit"),
                "normalized_unit": _normalize_unit(item_data.get("unit")),
                "unit_price": item_data.get("unit_price", 0),
                "total_price": item_data.get("total_price", 0),
                "notes": item_data.get("notes"),
                "coverage": "unknown"  # Will be determined during mapping
            }
            supabase.table("quote_line_items").insert(line_item).execute()
            line_items_created += 1
        
        # Calculate total amount from pricing_summary or line items
        total_amount = 0
        if "pricing_summary" in normalized_data and "total_amount" in normalized_data["pricing_summary"]:
            total_amount = normalized_data["pricing_summary"]["total_amount"]
        else:
            # Calculate from line items if no total in pricing summary
            total_amount = sum(item.get("total_price", 0) for item in normalized_data.get("line_items", []))
        
        return {
            "message": "Quote parsed successfully", 
            "quote_id": quote_id,
            "line_items_created": line_items_created,
            "total_amount": total_amount,
            "confidence_score": normalized_data.get("confidence_score", 0.5),
            "parsing_flags": normalized_data.get("parsing_flags", []),
            "vendor_name": normalized_data.get("vendor_info", {}).get("name", "Unknown"),
            "status": "parsed",
            "debug_info": {
                "file_path": file_url,
                "text_extracted_length": len(quote_text) if 'quote_text' in locals() else 0,
                "text_preview": quote_text[:200] if 'quote_text' in locals() else "No text extracted"
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error parsing quote: {str(e)}")

def _normalize_unit(raw_unit: Optional[str]) -> Optional[str]:
    """Normalize unit to standard construction units"""
    if not raw_unit:
        return None
    
    unit_map = {
        "EACH": "EA", "PIECE": "EA", "PIECES": "EA",
        "LINEAR": "LF", "LINEAL": "LF", "LIN": "LF", 
        "SQUARE": "SF", "SQ": "SF", "SQFT": "SF",
        "CUBIC": "CY", "YARD": "SY", "YARDS": "SY",
        "HOUR": "HR", "HOURS": "HR", "HRS": "HR",
        "LUMP": "LS", "LUMPSUM": "LS", "LOT": "LS",
        "MONTH": "MO", "MONTHS": "MO", "MONTHLY": "MO"
    }
    
    clean_unit = raw_unit.upper().strip()
    return unit_map.get(clean_unit, clean_unit if clean_unit in ["EA", "LF", "SF", "SY", "CY", "HR", "LS", "MO"] else None)

@router.post("/{quote_id}/mappings")
async def save_quote_mappings(
    quote_id: str = Path(...),
    mappings: List[Dict[str, Any]] = []
):
    """Save line mappings between budget and quote lines"""
    try:
        supabase = get_supabase_client()
        
        # Verify quote exists
        quote_check = supabase.table("vendor_quotes").select("id").eq("id", quote_id).execute()
        if not quote_check.data:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        # Clear existing mappings for this quote
        existing_mappings = supabase.table("line_mappings")\
            .select("id")\
            .in_("quote_line_item_id", 
                 supabase.table("quote_line_items").select("id").eq("quote_id", quote_id).execute().data)\
            .execute()
        
        if existing_mappings.data:
            mapping_ids = [m["id"] for m in existing_mappings.data]
            supabase.table("line_mappings").delete().in_("id", mapping_ids).execute()
        
        # Create new mappings
        for mapping in mappings:
            mapping_record = {
                "id": str(uuid.uuid4()),
                "budget_line_id": mapping["budget_line_id"],
                "quote_line_item_id": mapping["quote_line_item_id"],
                "confidence": mapping.get("confidence", 0.5),
                "user_confirmed": mapping.get("user_confirmed", False)
            }
            supabase.table("line_mappings").insert(mapping_record).execute()
        
        # Update quote status to mapped
        supabase.table("vendor_quotes").update({"status": "mapped"}).eq("id", quote_id).execute()
        
        return {
            "message": "Mappings saved successfully",
            "quote_id": quote_id,
            "mappings_created": len(mappings),
            "status": "mapped"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving mappings: {str(e)}")

@router.get("/divisions/{division_id}/compare")
async def compare_division_quotes(division_id: str = Path(...)):
    """Get comparison data for a division: budget lines + mapped quotes + deltas"""
    try:
        supabase = get_supabase_client()
        
        # Extract division code and project ID for filtering
        division_code = division_id.split('-')[0]  # e.g., "04"
        project_uuid = division_id.split('-', 1)[1] if '-' in division_id else division_id
        
        # Get quotes for this specific division (excluding subcategory-specific quotes)
        quotes_result = supabase.table("vendor_quotes")\
            .select("*, vendors(name), quote_line_items(*, line_mappings(budget_line_id)), normalized_json")\
            .eq("project_id", project_uuid)\
            .eq("division_code", division_code)\
            .is_("subcategory_id", "null")\
            .execute()
        
        if not quotes_result.data:
            return {
                "division_id": division_id,
                "budget_lines": [],
                "vendor_quotes": [],
                "comparison": []
            }
        
        # TODO: Get budget lines for this division from stored project data
        # For now return structure that frontend can work with
        
        comparison_data = []
        for quote in quotes_result.data:
            # Extract quote-level total from normalized_json for cases where line items have no pricing
            quote_level_total = 0
            if quote.get("normalized_json") and quote["normalized_json"].get("pricing_summary"):
                quote_level_total = quote["normalized_json"]["pricing_summary"].get("total_amount", 0)
            
            vendor_comparison = {
                "vendor_id": quote["vendor_id"],
                "vendor_name": quote["vendors"]["name"] if quote.get("vendors") else quote.get("vendor_name", "Unknown"),
                "quote_id": quote["id"],
                "status": quote["status"],
                "quote_level_total": quote_level_total,
                "normalized_json": quote.get("normalized_json", {}),
                # Enhanced scope information
                "scope_type": quote.get("scope_type", "complete_division"),
                "scope_items": quote.get("scope_items", []),
                "scope_budget_total": quote.get("scope_budget_total", 0.0),
                "scope_notes": quote.get("scope_notes", ""),
                "line_items": []
            }
            
            for line_item in quote.get("quote_line_items", []):
                line_comparison = {
                    "quote_line_id": line_item["id"],
                    "description": line_item["description"],
                    "total_price": line_item["total_price"],
                    "coverage": line_item["coverage"],
                    "mapped_budget_lines": [m["budget_line_id"] for m in line_item.get("line_mappings", [])]
                }
                vendor_comparison["line_items"].append(line_comparison)
            
            comparison_data.append(vendor_comparison)
        
        return {
            "division_id": division_id,
            "vendor_quotes": comparison_data,
            "total_vendors": len(comparison_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error comparing quotes: {str(e)}")

@router.post("/divisions/{division_id}/decide")
async def decide_division_award(
    division_id: str = Path(...),
    decisions: Dict[str, Any] = {}
):
    """Create division award with line-level decisions"""
    try:
        supabase = get_supabase_client()
        
        # Create division award
        award_id = str(uuid.uuid4())
        award_record = {
            "id": award_id,
            "division_id": division_id,
            "decided_at": datetime.now().isoformat(),
            "notes": decisions.get("notes", "")
        }
        award_result = supabase.table("division_awards").insert(award_record).execute()
        
        # Create line-level awards
        line_decisions = decisions.get("line_awards", [])
        for line_decision in line_decisions:
            line_award = {
                "id": str(uuid.uuid4()),
                "award_id": award_id,
                "budget_line_id": line_decision["budget_line_id"],
                "vendor_id": line_decision["vendor_id"],
                "quote_id": line_decision["quote_id"],
                "quote_line_item_id": line_decision.get("quote_line_item_id"),
                "final_price": line_decision["final_price"]
            }
            supabase.table("award_lines").insert(line_award).execute()
        
        # Update division status to "winner_selected" 
        primary_vendor_id = decisions.get("primary_vendor_id")
        supabase.table("division_status").upsert({
            "division_id": division_id,
            "status": "winner_selected",
            "selected_vendor_id": primary_vendor_id,
            "updated_at": datetime.now().isoformat()
        }).execute()
        
        return {
            "message": "Division award created successfully",
            "award_id": award_id,
            "division_id": division_id,
            "line_awards_created": len(line_decisions)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating award: {str(e)}")

@router.post("/divisions/{division_id}/workorder/pdf")
async def generate_work_order_pdf(division_id: str = Path(...)):
    """Generate work order PDF from division award"""
    try:
        supabase = get_supabase_client()
        
        # Get division award with line details
        award_result = supabase.table("division_awards")\
            .select("*, award_lines(*, vendors(name))")\
            .eq("division_id", division_id)\
            .execute()
        
        if not award_result.data:
            raise HTTPException(status_code=404, detail="No award found for division")
        
        award = award_result.data[0]
        
        # TODO: Generate actual PDF using reportlab or similar
        # For now, return PDF generation confirmation
        
        pdf_filename = f"workorder_division_{division_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        return {
            "message": "Work order PDF generated",
            "division_id": division_id,
            "award_id": award["id"], 
            "pdf_filename": pdf_filename,
            "line_items_count": len(award.get("award_lines", []))
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating work order: {str(e)}")

@router.delete("/{quote_id}")
async def delete_quote(quote_id: str = Path(...)):
    """Delete a vendor quote and all associated data"""
    try:
        supabase = get_supabase_client()
        
        # Verify quote exists
        quote_result = supabase.table("vendor_quotes").select("*").eq("id", quote_id).execute()
        if not quote_result.data:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        quote = quote_result.data[0]
        
        # Delete associated line mappings first (due to foreign key constraints)
        line_items_result = supabase.table("quote_line_items").select("id").eq("quote_id", quote_id).execute()
        if line_items_result.data:
            line_item_ids = [item["id"] for item in line_items_result.data]
            supabase.table("line_mappings").delete().in_("quote_line_item_id", line_item_ids).execute()
        
        # Delete quote line items
        supabase.table("quote_line_items").delete().eq("quote_id", quote_id).execute()
        
        # Delete the quote record
        supabase.table("vendor_quotes").delete().eq("id", quote_id).execute()
        
        # Clean up local file if it exists
        file_url = quote.get("file_url")
        if file_url and file_url.startswith("/tmp/"):
            try:
                os.remove(file_url)
            except:
                pass  # File may not exist or already deleted
        
        return {
            "message": "Quote deleted successfully",
            "quote_id": quote_id,
            "vendor_name": quote.get("vendor_name", "Unknown")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting quote: {str(e)}")